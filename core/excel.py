from xlrd import Book, open_workbook, xldate_as_tuple
from xlrd.sheet import Sheet
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Union, Any
from pathlib import Path
from datetime import datetime, date
import re
from abc import ABC, abstractmethod

re_sp = re.compile(r"\s+")
re_dt = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")


def _to_num(s: str):
    s = re.sub(r"\s*€$", "", s)
    if "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    f = float(s)
    i = int(f)
    return i if i == f else f


class Excel(ABC):
    @staticmethod
    def load(path: Union[str, Path]):
        if isinstance(path, str):
            path = Path(path)
        if path.suffix == ".xls":
            return Xls(path)
        if path.suffix == ".xlsx":
            return Xlsx(path)
        raise ValueError(f"Unsupported file extension: '{path.suffix}'")

    @abstractmethod
    def get(self, row: int, col: int) -> Any:
        ...

    @property
    def nrows(self) -> int:
        ...

    def get_text(self, row: int, col: int):
        s = self.get(row, col)
        if s is None:
            return None
        if not isinstance(s, str):
            raise TypeError(f"Expected str, got {type(s)}")
        s = re_sp.sub(" ", s)
        s = s.strip()
        if len(s) == 0:
            return None
        return s

    def get_word(self, row: int, col: int):
        s = self.get_text(row, col)
        if s is None:
            return None
        return re_sp.sub("", s)

    def get_number(self, row: int, col: int):
        s = self.get(row, col)
        if s is None:
            return None
        if isinstance(s, int):
            return s
        if isinstance(s, float):
            i = int(s)
            return i if i == s else s
        if not isinstance(s, str):
            raise TypeError(f"Expected int|foat, got {type(s)}")
        s = self.get_text(row, col)
        if s is None:
            return None
        return _to_num(s)

    def get_date(self, row: int, col: int):
        s = self.get(row, col)
        if s is None:
            return None
        if isinstance(s, datetime):
            return s.date()
        if isinstance(s, date):
            return s
        if isinstance(s, str):
            m = re_dt.match(s)
            if s:
                return date(*map(int, reversed(m.groups())))
        raise TypeError(f"Expected date, got {type(s)}")


class Xls(Excel):
    def __init__(self, path: Union[str, Path]):
        if isinstance(path, str):
            path = Path(path)
        if not path.is_file():
            raise FileNotFoundError(f"No such file: '{path}'")
        self.wb: Book = open_workbook(path)
        self.ws: Sheet = self.wb.sheet_by_index(0)

    def get(self, row: int, col: int):
        return self.ws.cell(row, col).value

    def get_date(self, row: int, col: int):
        s = self.get(row, col)
        if s is None:
            return None
        d = datetime(*xldate_as_tuple(s, self.wb.datemode))
        return d.date()

    @property
    def nrows(self):
        return self.ws.nrows


class Xlsx(Excel):
    def __init__(self, path: Union[str, Path]):
        self.wb = load_workbook(path, data_only=True)
        self.ws: Worksheet = self.wb.active

    def get(self, row: int, col: int):
        return self.ws.cell(row=row+1, column=col+1).value

    @property
    def nrows(self):
        return self.ws.max_row
